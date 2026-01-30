import re
import pandas as pd
from datetime import datetime
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters
)

import os
TOKEN = os.getenv("TOKEN")

buffer = []

def parse_load(text: str):
    try:
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        # Load ID
        load_id = lines[0]

        # Rate — первая строка с $
        rate = None
        for l in lines:
            if "$" in l:
                rate_str = l.replace("$", "").strip()
                rate_str = rate_str.replace(" ", "")

                if rate_str.count(",") == 1 and rate_str.count(".") == 1:
                    rate_str = rate_str.replace(",", "")
                else:
                    rate_str = rate_str.replace(",", ".")

                rate = float(rate_str)
                break
        if rate is None:
            return None

        # Города — любая строка с запятой
        cities = []
        for l in lines:
            if "," in l:
                city = l.split(",")[0].strip()
                if len(city) > 2 and not any(c.isdigit() for c in city):
                    cities.append(city)

        if len(cities) < 2:
            return None

        pickup_city = cities[0]
        delivery_city = cities[-1]

        # Delivery date — последняя дата
        delivery_date = None
        for l in reversed(lines):
            m = re.search(r"(\d{1,2})\s+([A-Za-z]{3})", l)
            if m:
                delivery_date = datetime.strptime(
                    f"{m.group(1)} {m.group(2)} 2026",
                    "%d %b %Y"
                ).strftime("%m/%d/%Y")
                break
        if delivery_date is None:
            return None

        return {
            "Load ID": load_id,
            "Route": f"{pickup_city} - {delivery_city}",
            "Rate": rate,
            "Broker": "Amazon",
            "Date": delivery_date,
            "Status": "completed"
        }

    except:
        return None


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buffer.clear()
    await update.message.reply_text(
        "🟢 Готов.\n"
        "Кидай грузы.\n"
        "Потом /excel"
    )


async def make_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not buffer:
        await update.message.reply_text("❌ Нет данных")
        return

    df = pd.DataFrame(buffer)

    filename = f"loads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    df.to_excel(filename, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

    wb = load_workbook(filename)
    ws = wb.active

    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")

    # Заголовки
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center

    # Тело таблицы
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = center

    # Формат чисел
    for cell in ws["C"][1:]:  # Rate
        cell.number_format = "0.00"

    for cell in ws["E"][1:]:  # Date
        cell.number_format = "M/D/YYYY"

    # Авто-ширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)

    await update.message.reply_document(open(filename, "rb"))
    buffer.clear()


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parsed = parse_load(update.message.text)
    if parsed:
        buffer.append(parsed)
        await update.message.reply_text("✅ Груз добавлен")
    else:
        await update.message.reply_text("⚠️ Формат не распознан")


def main():
    print("DEBUG TOKEN =", repr(TOKEN))
    print("DEBUG LEN =", len(TOKEN) if TOKEN else None)

    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("excel", make_excel))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    app.run_polling()


if __name__ == "__main__":
    main()
